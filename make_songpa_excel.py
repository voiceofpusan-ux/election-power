"""
송파구 오프라인 투표 집계 엑셀 생성기
시트1: 갑/을/병 통합 집계표 (직접 입력)
시트2: 막대 그래프
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation

WB_PATH = r"e:\github\election-power\송파구_투표집계.xlsx"

DONGS = {
    "갑": ["풍납1동","풍납2동","방이1동","방이2동","오륜동","송파1동","송파2동","잠실4동","잠실6동"],
    "을": ["잠실본동","잠실2동","잠실3동","잠실7동","석촌동","삼전동","가락1동","문정2동"],
    "병": ["거여1동","거여2동","마천1동","마천2동","오금동","가락본동","가락2동","문정1동","장지동","위례동"],
}

# ── 색상 팔레트 ──────────────────────────────────────
C_TITLE_BG   = "1a56db"   # 진파랑 (제목)
C_GAP_BG     = "1e429f"   # 갑 헤더
C_EUL_BG     = "7e3af2"   # 을 헤더
C_BYEONG_BG  = "0e9f6e"   # 병 헤더
C_SUBTOTAL   = "e1effe"   # 소계 행 연하늘
C_TOTAL_BG   = "1a56db"   # 합계 행
C_INPUT_FILL = "fffde7"   # 입력 셀 연노랑
C_DANG_HDR   = "fef3c7"   # 당원 헤더 노랑
C_ILBAN_HDR  = "ecfdf5"   # 일반 헤더 민트
C_WHITE      = "FFFFFF"
C_GRAY       = "f3f4f6"
C_BORDER     = "9ca3af"

UNIT_COLORS = {"갑": C_GAP_BG, "을": C_EUL_BG, "병": C_BYEONG_BG}
UNIT_KR     = {"갑": "갑", "을": "을", "병": "병"}

def side(style="thin", color=C_BORDER):
    return Side(style=style, color=color)

def border_all(thick=False):
    s = "medium" if thick else "thin"
    return Border(left=side(s), right=side(s), top=side(s), bottom=side(s))

def border_thick():
    return Border(
        left=side("medium"), right=side("medium"),
        top=side("medium"), bottom=side("medium")
    )

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, white=False):
    c = "FFFFFF" if white else color
    return Font(bold=bold, color=c, size=size, name="맑은 고딕")

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

# ── 컬럼 레이아웃 ─────────────────────────────────────
# A   B       C       D       E       F
# 단위 동명   당원    일반    합계    비고
COL_UNIT  = 1   # A
COL_DONG  = 2   # B
COL_DANG  = 3   # C
COL_ILBAN = 4   # D
COL_SUM   = 5   # E
COL_NOTE  = 6   # F

def make_sheet1(wb):
    ws = wb.active
    ws.title = "송파구 집계"
    ws.sheet_view.showGridLines = False

    # ── 열 너비 ───────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16

    # ── 행 1~2: 제목 ──────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "송파구 투표 집계표"
    c.font = font(bold=True, size=16, white=True)
    c.fill = fill(C_TITLE_BG)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "※ 당원투표(C열)와 일반(안심번호)(D열)에 숫자를 직접 입력하세요. 합계는 자동 계산됩니다."
    c.font = Font(size=9, color="4b5563", name="맑은 고딕", italic=True)
    c.fill = fill("eff6ff")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # ── 행 3: 컬럼 헤더 ──────────────────────────────
    headers = ["단위", "동명", "당원투표", "일반(안심번호)", "합계", "비고"]
    hdr_fills = [C_TITLE_BG, C_TITLE_BG, "b45309", "065f46", C_TITLE_BG, C_TITLE_BG]
    for col_idx, (h, bg) in enumerate(zip(headers, hdr_fills), start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = font(bold=True, size=10, white=True)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border_all()
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    current_row = 4
    subtotal_refs = {}   # unit → (dang_col_letter, row) for chart

    for unit in ["갑", "을", "병"]:
        dongs = DONGS[unit]
        uc = UNIT_COLORS[unit]
        start_data_row = current_row

        # ── 단위 헤더 행 ──────────────────────────────
        ws.merge_cells(f"A{current_row}:F{current_row}")
        c = ws.cell(row=current_row, column=1)
        c.value = f"▶  {unit}선거구  ({len(dongs)}개 동)"
        c.font = font(bold=True, size=11, white=True)
        c.fill = fill(uc)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border = border_thick()
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        dong_rows = []

        for dong in dongs:
            r = current_row
            dong_rows.append(r)
            ws.row_dimensions[r].height = 20

            # 단위 셀 (병합 없이 값만, 단위 헤더 아래 첫 행만 표시)
            c_unit = ws.cell(row=r, column=COL_UNIT, value="" )
            c_unit.fill = fill(C_GRAY)
            c_unit.alignment = center()
            c_unit.border = border_all()

            # 동명
            c_dong = ws.cell(row=r, column=COL_DONG, value=dong)
            c_dong.font = font(size=10)
            c_dong.fill = fill(C_WHITE)
            c_dong.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c_dong.border = border_all()

            # 당원투표 입력
            c_dang = ws.cell(row=r, column=COL_DANG, value=0)
            c_dang.font = font(size=10)
            c_dang.fill = fill(C_INPUT_FILL)
            c_dang.alignment = center()
            c_dang.border = border_all()
            c_dang.number_format = "#,##0"

            # 일반(안심번호) 입력
            c_ilban = ws.cell(row=r, column=COL_ILBAN, value=0)
            c_ilban.font = font(size=10)
            c_ilban.fill = fill(C_INPUT_FILL)
            c_ilban.alignment = center()
            c_ilban.border = border_all()
            c_ilban.number_format = "#,##0"

            # 합계 수식
            c_sum = ws.cell(row=r, column=COL_SUM)
            c_sum.value = f"=C{r}+D{r}"
            c_sum.font = font(bold=True, size=10)
            c_sum.fill = fill("f0f9ff")
            c_sum.alignment = center()
            c_sum.border = border_all()
            c_sum.number_format = "#,##0"

            # 비고
            c_note = ws.cell(row=r, column=COL_NOTE, value="")
            c_note.fill = fill(C_WHITE)
            c_note.border = border_all()

            current_row += 1

        # ── 소계 행 ───────────────────────────────────
        r = current_row
        ws.row_dimensions[r].height = 22

        c_label = ws.cell(row=r, column=COL_UNIT)
        c_label.fill = fill(C_SUBTOTAL)
        c_label.border = border_all()

        ws.merge_cells(f"B{r}:B{r}")
        c_dong = ws.cell(row=r, column=COL_DONG, value=f"{unit} 소계")
        c_dong.font = font(bold=True, size=10, color="1e3a5f")
        c_dong.fill = fill(C_SUBTOTAL)
        c_dong.alignment = center()
        c_dong.border = border_all()

        dang_sum_range  = f"C{start_data_row}:C{current_row-1}"
        ilban_sum_range = f"D{start_data_row}:D{current_row-1}"

        c_dang_sub = ws.cell(row=r, column=COL_DANG, value=f"=SUM({dang_sum_range})")
        c_dang_sub.font = font(bold=True, size=10, color="92400e")
        c_dang_sub.fill = fill(C_SUBTOTAL)
        c_dang_sub.alignment = center()
        c_dang_sub.border = border_all()
        c_dang_sub.number_format = "#,##0"

        c_ilban_sub = ws.cell(row=r, column=COL_ILBAN, value=f"=SUM({ilban_sum_range})")
        c_ilban_sub.font = font(bold=True, size=10, color="065f46")
        c_ilban_sub.fill = fill(C_SUBTOTAL)
        c_ilban_sub.alignment = center()
        c_ilban_sub.border = border_all()
        c_ilban_sub.number_format = "#,##0"

        c_total_sub = ws.cell(row=r, column=COL_SUM, value=f"=C{r}+D{r}")
        c_total_sub.font = font(bold=True, size=10, color="1e3a5f")
        c_total_sub.fill = fill(C_SUBTOTAL)
        c_total_sub.alignment = center()
        c_total_sub.border = border_all()
        c_total_sub.number_format = "#,##0"

        ws.cell(row=r, column=COL_NOTE).border = border_all()
        ws.cell(row=r, column=COL_NOTE).fill = fill(C_SUBTOTAL)

        subtotal_refs[unit] = r
        current_row += 1

    # ── 최종 합계 행 ──────────────────────────────────
    r = current_row
    ws.row_dimensions[r].height = 26

    ws.merge_cells(f"A{r}:B{r}")
    c_label = ws.cell(row=r, column=1, value="송파구 전체 합계")
    c_label.font = font(bold=True, size=12, white=True)
    c_label.fill = fill(C_TOTAL_BG)
    c_label.alignment = center()
    c_label.border = border_thick()

    gap_r  = subtotal_refs["갑"]
    eul_r  = subtotal_refs["을"]
    by_r   = subtotal_refs["병"]

    c_dang_total = ws.cell(row=r, column=COL_DANG,
                           value=f"=C{gap_r}+C{eul_r}+C{by_r}")
    c_dang_total.font = font(bold=True, size=12, white=True)
    c_dang_total.fill = fill(C_TOTAL_BG)
    c_dang_total.alignment = center()
    c_dang_total.border = border_thick()
    c_dang_total.number_format = "#,##0"

    c_ilban_total = ws.cell(row=r, column=COL_ILBAN,
                            value=f"=D{gap_r}+D{eul_r}+D{by_r}")
    c_ilban_total.font = font(bold=True, size=12, white=True)
    c_ilban_total.fill = fill(C_TOTAL_BG)
    c_ilban_total.alignment = center()
    c_ilban_total.border = border_thick()
    c_ilban_total.number_format = "#,##0"

    c_sum_total = ws.cell(row=r, column=COL_SUM,
                          value=f"=C{r}+D{r}")
    c_sum_total.font = font(bold=True, size=12, white=True)
    c_sum_total.fill = fill(C_TOTAL_BG)
    c_sum_total.alignment = center()
    c_sum_total.border = border_thick()
    c_sum_total.number_format = "#,##0"

    ws.cell(row=r, column=COL_NOTE).fill = fill(C_TOTAL_BG)
    ws.cell(row=r, column=COL_NOTE).border = border_thick()

    return ws, subtotal_refs, current_row


def make_sheet2(wb, subtotal_refs):
    """차트 시트: 갑/을/병 단위별 당원 vs 일반 막대그래프 + 동별 세부 차트"""
    ws = wb.create_sheet("차트")
    ws.sheet_view.showGridLines = False
    ws1 = wb["송파구 집계"]

    # ── 차트 데이터 테이블 (숨겨진 보조 테이블) ─────────
    # H열부터 차트용 데이터 기록
    ws.column_dimensions["A"].width = 2

    # 소계 차트 데이터
    ws["B2"] = "구분"
    ws["C2"] = "당원투표"
    ws["D2"] = "일반(안심번호)"
    ws["E2"] = "합계"
    for col in ["B","C","D","E"]:
        c = ws[f"{col}2"]
        c.font = font(bold=True, white=True)
        c.fill = fill(C_TITLE_BG)
        c.alignment = center()
        c.border = border_all()

    unit_map = {"갑": 3, "을": 4, "병": 5}
    for unit, row in unit_map.items():
        r = subtotal_refs[unit]
        ws.cell(row=row, column=2, value=f"{unit}선거구").font = font(bold=True)
        ws.cell(row=row, column=2).alignment = center()
        ws.cell(row=row, column=2).border = border_all()

        for col_dst, col_src in [(3, "C"), (4, "D"), (5, "E")]:
            c = ws.cell(row=row, column=col_dst)
            c.value = f"='{ws1.title}'!{col_src}{r}"
            c.number_format = "#,##0"
            c.alignment = center()
            c.border = border_all()

    # ── 차트 1: 단위별 당원 vs 일반 묶음 막대 ──────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "clustered"
    chart1.title = "송파구 선거구별 투표 현황"
    chart1.style = 10
    chart1.y_axis.title = "투표수"
    chart1.x_axis.title = "선거구"
    chart1.y_axis.numFmt = "#,##0"
    chart1.width  = 20
    chart1.height = 14

    cats = Reference(ws, min_col=2, min_row=3, max_row=5)
    data_dang  = Reference(ws, min_col=3, min_row=2, max_row=5)
    data_ilban = Reference(ws, min_col=4, min_row=2, max_row=5)

    chart1.add_data(data_dang,  titles_from_data=True)
    chart1.add_data(data_ilban, titles_from_data=True)
    chart1.set_categories(cats)

    # 당원=파랑, 일반=초록
    chart1.series[0].graphicalProperties.solidFill = "1a56db"
    chart1.series[0].graphicalProperties.line.solidFill = "1a56db"
    chart1.series[1].graphicalProperties.solidFill = "0e9f6e"
    chart1.series[1].graphicalProperties.line.solidFill = "0e9f6e"

    ws.add_chart(chart1, "B7")

    # ── 차트 2: 단위별 합계 누적 막대 ────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "선거구별 누적 합계"
    chart2.style = 10
    chart2.y_axis.title = "투표수"
    chart2.x_axis.title = "선거구"
    chart2.y_axis.numFmt = "#,##0"
    chart2.width  = 14
    chart2.height = 14

    data_total = Reference(ws, min_col=5, min_row=2, max_row=5)
    chart2.add_data(data_total, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.solidFill = "7e3af2"

    ws.add_chart(chart2, "M7")

    # ── 동별 세부 데이터 테이블 (B28부터) ────────────────
    start_row = 28
    ws.cell(row=start_row-1, column=2, value="▶ 동별 세부 현황 (차트 원본 데이터)").font = font(bold=True, size=11)
    headers = ["단위","동명","당원투표","일반(안심번호)","합계"]
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = font(bold=True, white=True)
        c.fill = fill(C_TITLE_BG)
        c.alignment = center()
        c.border = border_all()

    data_row = start_row + 1
    # 동 데이터를 시트1에서 참조
    # 시트1의 행 위치를 재계산
    r1_row = 4  # 시트1 데이터 시작
    for unit in ["갑","을","병"]:
        r1_row += 1  # 단위 헤더
        for dong in DONGS[unit]:
            ws.cell(row=data_row, column=2, value=unit).alignment = center()
            ws.cell(row=data_row, column=2).border = border_all()
            ws.cell(row=data_row, column=3, value=dong).border = border_all()
            ws.cell(row=data_row, column=3).alignment = Alignment(horizontal="left", indent=1)

            for col_dst, col_src in [(4,"C"),(5,"D"),(6,"E")]:
                c = ws.cell(row=data_row, column=col_dst)
                c.value = f"='{ws1.title}'!{col_src}{r1_row}"
                c.number_format = "#,##0"
                c.alignment = center()
                c.border = border_all()

            r1_row += 1
            data_row += 1
        r1_row += 1  # 소계 행

    # ── 차트 3: 동별 세부 막대 ────────────────────────────
    chart3 = BarChart()
    chart3.type = "bar"   # 가로 막대 (동 이름이 길어서)
    chart3.grouping = "clustered"
    chart3.title = "동별 투표 현황 (당원 vs 일반)"
    chart3.style = 10
    chart3.x_axis.title = "투표수"
    chart3.y_axis.numFmt = "#,##0"
    chart3.width  = 28
    chart3.height = 30

    last_data_row = data_row - 1
    cats3  = Reference(ws, min_col=3, min_row=start_row+1, max_row=last_data_row)
    dang3  = Reference(ws, min_col=4, min_row=start_row,   max_row=last_data_row)
    ilban3 = Reference(ws, min_col=5, min_row=start_row,   max_row=last_data_row)

    chart3.add_data(dang3,  titles_from_data=True)
    chart3.add_data(ilban3, titles_from_data=True)
    chart3.set_categories(cats3)

    chart3.series[0].graphicalProperties.solidFill = "1a56db"
    chart3.series[0].graphicalProperties.line.solidFill = "1a56db"
    chart3.series[1].graphicalProperties.solidFill = "0e9f6e"
    chart3.series[1].graphicalProperties.line.solidFill = "0e9f6e"

    ws.add_chart(chart3, "B31")

    return ws


def main():
    wb = openpyxl.Workbook()
    ws1, subtotal_refs, _ = make_sheet1(wb)
    make_sheet2(wb, subtotal_refs)

    wb.active = wb["송파구 집계"]
    wb.save(WB_PATH)
    print(f"저장 완료: {WB_PATH}")


if __name__ == "__main__":
    main()
